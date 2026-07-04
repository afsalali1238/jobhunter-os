#!/usr/bin/env python3
"""
Deterministic helper for JobHunter OS: append or update a lead row in
excel/JobHunter_Pipeline.xlsx's Pipeline table, without hand-rolling openpyxl
code from scratch each time (which is error-prone re: table ranges,
conditional formatting, and band text).

Usage:
  python3 add_lead.py --file excel/JobHunter_Pipeline.xlsx \
      --company "Acme Corp" --role "Ops Manager" --url "https://..." \
      [--score 82] [--status Scouted] [--applied-date "04 Jul"]

Behavior:
  - If a row already has this URL (normalized: trimmed, lowercased, no
    query string/fragment), that row is UPDATED in place (used by score-fit
    when re-scoring an existing lead) instead of creating a duplicate.
  - Otherwise a new row is appended right after the last used row.
  - The "Band" column is plain text (HOT/Strong/Consider/Low), computed
    from --score using the same cutoffs as AGENTS.md. This sheet has no
    Band formula — conditional formatting colors it based on that text,
    already pre-applied to D2:D500. Leave --score blank to leave Band blank.
  - The PipelineTable range is expanded to cover the new last row so
    Excel's filters/banding treat it as part of the table.

Requires: pip install openpyxl --break-system-packages (if not already available).
"""
import argparse
import sys
from urllib.parse import urlsplit, urlunsplit

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is not installed. Run: pip install openpyxl --break-system-packages", file=sys.stderr)
    sys.exit(1)


def band_for(score):
    if score is None or score == "":
        return ""
    score = int(score)
    if score >= 80:
        return "HOT"
    if score >= 70:
        return "Strong"
    if score >= 60:
        return "Consider"
    return "Low"


def normalize_url(url):
    """Strip query string/fragment, trailing slash, lowercase, trim — so the
    same posting re-listed with a tracking param (?utm_source=...) is still
    recognized as the same lead. Mirrors dashboard/app.js's normalizeUrl()."""
    if not url:
        return ""
    url = url.strip().lower()
    try:
        parts = urlsplit(url)
        cleaned = urlunsplit((parts.scheme, parts.netloc, parts.path.rstrip('/'), '', ''))
        return cleaned
    except Exception:
        return url


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True)
    ap.add_argument("--company", required=True)
    ap.add_argument("--role", required=True)
    ap.add_argument("--url", required=True)
    ap.add_argument("--score", default=None)
    ap.add_argument("--status", default="Scouted")
    ap.add_argument("--applied-date", default="")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.file)
    if "Pipeline" not in wb.sheetnames:
        print("ERROR: no 'Pipeline' sheet found in this workbook.", file=sys.stderr)
        sys.exit(1)
    ws = wb["Pipeline"]

    band = band_for(args.score)
    target_url_norm = normalize_url(args.url)

    # Scan every used row (don't stop early on a match) so we get both the
    # correct existing_row (if this lead is already there) AND the true
    # last used row, needed to know where to append if it's new.
    existing_row = None
    row_idx = 2
    last_used_row = 1
    while ws.cell(row=row_idx, column=1).value is not None:
        existing_url = ws.cell(row=row_idx, column=7).value
        if target_url_norm and existing_row is None and normalize_url(existing_url) == target_url_norm:
            existing_row = row_idx
        last_used_row = row_idx
        row_idx += 1

    if existing_row:
        target_row = existing_row
        action = "updated"
    else:
        target_row = last_used_row + 1
        action = "appended"

    ws.cell(row=target_row, column=1, value=args.company)
    ws.cell(row=target_row, column=2, value=args.role)
    ws.cell(row=target_row, column=3, value=int(args.score) if args.score not in (None, "") else None)
    ws.cell(row=target_row, column=4, value=band if band else None)
    ws.cell(row=target_row, column=5, value=args.status)
    ws.cell(row=target_row, column=6, value=args.applied_date if args.applied_date else None)
    ws.cell(row=target_row, column=7, value=args.url)

    # Expand the PipelineTable range so Excel's filter/banding covers the new row.
    new_last_row = max(target_row, last_used_row)
    if "PipelineTable" in ws.tables:
        tbl = ws.tables["PipelineTable"]
        tbl.ref = f"A1:G{new_last_row}"

    wb.save(args.file)
    print(f"{action.capitalize()} row {target_row}: {args.company} / {args.role} (score={args.score or 'none'}, band={band or 'none'})")


if __name__ == "__main__":
    main()
